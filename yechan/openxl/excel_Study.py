#-*- coding: utf-8 -*-
import openpyxl
import sys
reload(sys)
sys.setdefaultencoding('utf-8')



# open excel file load_workbook 으로 읽어들인 객체의 타입은 workbook 객체
document = openpyxl.load_workbook('sample.xlsx')

print type(document)


# check sheets' names 

print document.get_sheet_names()
print type(document.get_sheet_names())

sheet = document.get_sheet_by_name("Sheet1")

print type(sheet)

print sheet['A2'].value

print type(sheet['A2']),type(sheet['A2'].value)

print sheet.cell(row=5,column=2).value # 다른 셀값 표현 방식

multplecells =sheet['A1':'B3']

for row in multplecells:
    for cell in row:
        print cell.value
#---------------------------------------------------------------------------------------------------------------------
# 워크북 생성 및 수정
# 타입은 튜플
all_rows =sheet.rows
print all_rows
# 워크북 생성
wb =Workbook()
ws1 = wb.active
ws1.title="first_sheet"
ws1["A1"]= "hello world"
print(ws1["A1"].value)

wb.save("C:/excel/test.xlsx")

wb2 = openpyxl.load_workbook("file.xlsx")

wb2.create_sheet(3,title="sheet name")
for row in range(1,10):
    for col in range(1,10):
        ws2.cell(row = row, column = col, value =int ("{}{}".format(row,col)))

# 셀 병항

ws.merge_cells("A1:B2")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
wb.save()

#셀  색칠
from openpyxl.styles import Patternfill

db_sheet['B' + str(db_i + 5)].fill = PatternFill("solid", fgColor="77FF77")