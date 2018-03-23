#-*-coding:utf-8-*-
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb =Workbook()

filename='newbook3.xlsx'
ws1 = wb.active #u워크 시트 생성
ws1.title ="range names" #u타이틀 생성

for row in range(1,40):
    ws1.append(range(600))
#u워크 북은 무슨 형식이지? append 가능 어떤range?
ws2 =wb.create_sheet(title="pi") #두번째 부턴 create_sheet()써야 한다
ws2['F5'] =3.14
# 왜 처음 건 ative로 하고 마지막에 이름을 save를 하는거지?
ws3 =wb.create_sheet(title="data")
for row in range(10,20):
    for col in range (27,54):
        _=ws3.cell(row=row, coulumn =col, value="{}".format(get_column_letter(col)))
print "ws4 ws5 연습"
ws4 =wb.create_sheet(title ="list insert")
ws5 = wb.create_sheet(title="tuple insert")
ws4.append([1,2,34])
ws5.append((1,2,34))
print ws4['A1'].value
ws4['A1'].value=12
print ws4['A1'].value
print '--------------'
print type(ws5.rows),type(ws4.rows), type(ws5['A1']),type(ws4['A1']) 
print ws5['A1'].value
print (ws3['AA10'].value)
ws6 =wb.create_sheet()
print ws6.title
wb.save(filename)
