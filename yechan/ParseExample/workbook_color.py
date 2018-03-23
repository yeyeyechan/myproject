#-*- coding: utf-8 -*-
import sys
import random
reload(sys)
sys.setdefaultencoding('utf-8')
from openpyxl.compat import range
from openpyxl import Workbook

from openpyxl.styles import PatternFill, Color
wb= Workbook()
ws1= wb.active

for row in range(1,40):
    for col in range(1,40):
        ws1.cell(column=col, row =row , value= random.randrange(1,20))

ws1.title ="color practice"

for row in range(1,40):
    for col in range(1,40):
        if ws1.cell(row =row, column=col).value >=10 :
            ws1.cell(row=row, column=col).fill = PatternFill(patternType ='solid', fgColor=Color('FFC000') )

wb.save("colorpractice.xlsx")





