# coding=utf-8
import datetime
from openpyxl.styles import PatternFill
import openpyxl
import argparse
import sys


def parse_args():
    story = u"DB 파일명, 데이터 시작셀과 마지막셀 입력"
    usg = u'\n python test.py -db "데이터보고서" -ds B5 -de B110'
    parser = argparse.ArgumentParser(description=story, usage=usg, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument("-db", default=None, help=u"DB 파일명, e.g.) DB검색_2017_0703_17")
    parser.add_argument("-ds", default=None, help=u"DB 데이터 시작 셀, e.g.) B5")
    parser.add_argument("-de", default=None, help=u"DB 데이터 마지막 셀, e.g.) B1100")
    parser.add_argument("-list", default=u"_2016년 사업장 별 사전계측용량(kW) 리스트", help=u"사업장 리스트 파일명, 기본값: _2016년 사업장 별 사전계측용량(kW) 리스트")
    parser.add_argument("-ls", default='A2', help=u"사업장 데이터 시작 셀, 기본값: A2")
    parser.add_argument("-le", default='A580', help=u"사업장 데이터 마지막 셀, 기본값: A580")

    args = parser.parse_args()

    # check
    _lst = args.list
    ls=args.ls
    le=args.le

    db = args.db
    if db == None:
        exit(u"DB 파일명 입력해주세요.")
    ds = args.ds
    if ds == None:
        exit(u"DB 데이터 시작 셀 입력해주세요")
    de = args.de
    if de == None:
        exit(u"DB 데이터 마지막 셀 입력해주세요")

    return _lst, db, ls, le, ds, de


# from "sample.py"
#### Author : jeonghoonkang , https://github.com/jeonghoonkang
####  Author : jeongmoon417 , https://github.com/jeongmoon417
class excell_class :
    __ofile = None

    def __init__(self):
        pass

    #@staticmethod
    def open_exc_doc(self,__file):
        # using unicode file name with u syntax
        __ofile = openpyxl.load_workbook(__file)

        # pout = "ix: %s, mdsid: %s, donecheck: %d \r" %(ix, mds_id, donecheck)
        pout = "   ... file opened \n"
        sys.stdout.write(pout)
        sys.stdout.flush()
        return __ofile

    def read_vertical(self, sheet, __start, __end):
        __vertical = []
        # print " ... Please use column[n]:column[m], vertical read "
        cell_of_col = sheet[__start:__end]
        for row in cell_of_col:
            for cell in row:
                v = cell.value
                if v == None:
                    continue
                __vertical.append(v)
        return __vertical  # __cnt, __cnt_n # 세로 셀 데이터, 데이터 갯수, None 갯수


def check_form(__buf):
    __itter = len(__buf)
    __list = []

    for __i in range(__itter) :
        if str(__buf[__i])[0:2] != '20':
            __buf[__i]='20'+str(__buf[__i])

        if len(str(__buf[__i])) == 12:
            __buf[__i] = str(__buf[__i])+'-01'

        elif len(str(__buf[__i])) > 12:
            if str(__buf[__i])[12] == '-':
                if str(__buf[__i])[13] != '0':
                    temp=__buf[__i][13]
                    __buf[__i]=str(__buf[__i])[0:12]+'-0'+temp
                    #print __buf[__i]+'t'

        #print __buf[__i]
        __list.append(__buf[__i])

    return __list


def color_list(__class, target, db):
    def __init__():
        pass

    # load db file
    db_sheets = db.get_sheet_names()
    db_sheet = db.get_sheet_by_name(db_sheets[0])
    db_company_nums = __class.read_vertical(db_sheet, db_start_range, db_end_range)

    # load list file
    target_sheets = target.get_sheet_names()
    target_sheet = target.get_sheet_by_name(target_sheets[0])
    target_company_nums = __class.read_vertical(target_sheet, list_start_range, list_end_range)

    check_form(db_company_nums)
    check_form(target_company_nums)

    # db data index
    dd = len(db_company_nums)
    # target data index
    td = len(target_company_nums)

    count_date = 0
    count_12digits = 0
    count_all = 0

    # compare and color - 8 digits(date) are the same
    for target_i in range(0, td):
        for db_i in range(0, dd):
            if str(db_company_nums[db_i])[0:8] == str(target_company_nums[target_i])[0:8]:
                target_sheet['A' + str(target_i+int(list_start_range[1:]))].fill = PatternFill("solid", fgColor="DDFFDD")
                count_date = count_date + 1
                # print str(db_company_nums[db_i])[0:8]
                break
    target_sheet['F2'].fill = PatternFill("solid", fgColor="DDFFDD")

    # compare and color - 12 digits are the same
    for target_i in range(0, td):
        for db_i in range(0, dd):
            if str(db_company_nums[db_i])[0:12] == str(target_company_nums[target_i])[0:12]:
                target_sheet['A' + str(target_i+int(list_start_range[1:]))].fill = PatternFill("solid", fgColor="77FF77")
                count_12digits = count_12digits + 1
                # print str(db_company_nums[db_i])[0:12]
                break
    target_sheet['E2'].fill = PatternFill("solid", fgColor="77FF77")


    # compare and color - exactly the same
    for target_i in range(0, td):
        for db_i in range(0, dd):
            if str(db_company_nums[db_i]) == str(target_company_nums[target_i]):
                target_sheet['A' + str(target_i+int(list_start_range[1:]))].fill = PatternFill("solid", fgColor="00BB00")
                count_all = count_all + 1
                # print str(db_company_nums[db_i])
                break
    target_sheet['D2'].fill = PatternFill("solid", fgColor="00BB00")

    target_sheet['D2'] = count_all
    target_sheet['E2'] = count_12digits - count_all
    target_sheet['F2'] = count_date - count_12digits

    # save colored copy file
    __t = datetime.datetime.now().strftime("%Y-%m-%d %H.%M.%S")
    file_name = __t + '_colored_result_for_' + db_file + '.xlsx'
    target.save(file_name)


def color_db(__class, company, db):
    def __init__():
        pass

    # load db file
    db_sheets = db.get_sheet_names()
    db_sheet = db.get_sheet_by_name(db_sheets[0])
    db_company_nums = __class.read_vertical(db_sheet, 'b5', 'b1104')

    # load list file
    company_sheets = company.get_sheet_names()
    company_sheet = company.get_sheet_by_name(company_sheets[0])
    company_company_nums = __class.read_vertical(company_sheet, 'a2', 'a580')

    check_form(db_company_nums)
    check_form(company_company_nums)

    # db data index
    dd = len(db_company_nums)
    # company data index
    td = len(company_company_nums)

    # compare and color - 8 digits(date) are the same
    for db_i in range(0, dd):
        for company_i in range(0, td):
            if str(db_company_nums[db_i])[0:8] == str(company_company_nums[company_i])[0:8]:
                db_sheet['B' + str(db_i + 5)].fill = PatternFill("solid", fgColor="DDFFDD")
                # print str(db_company_nums[db_i])[0:8]

    # compare and color - 12 digits are the same
    for db_i in range(0, dd):
        for company_i in range(0, td):
            if str(db_company_nums[db_i])[0:12] == str(company_company_nums[company_i])[0:12]:
                db_sheet['B' + str(db_i + 5)].fill = PatternFill("solid", fgColor="77FF77")
                # print str(db_company_nums[db_i])[0:12]

    # compare and color - exactly the same
    for db_i in range(0, dd):
        for company_i in range(0, td):
            if str(db_company_nums[db_i]) == str(company_company_nums[company_i]):
                db_sheet['B' + str(db_i + 5)].fill = PatternFill("solid", fgColor="00BB00")
                # print str(db_company_nums[db_i])


    # save colored copy file
    __t = datetime.datetime.now().strftime("%Y-%m-%d %H.%M.%S")
    file_name = __t + '_colored_result' + db_file + '.xlsx'
    db.save(file_name)


if __name__ == "__main__":
    list_file, db_file, list_start_range, list_end_range, db_start_range, db_end_range = parse_args()

    # 입력파일
    # LIST_file =  u"_2016년 사업장 별 사전계측용량(kW) 리스트.xlsx"
    # DB_file =  u"DB검색_2017_0703_17.xlsx"

    cfile = excell_class()
    _list = cfile.open_exc_doc(list_file + '.xlsx')
    _db = cfile.open_exc_doc(db_file + '.xlsx')

    color_list(cfile, _list, _db)
    color_db(cfile, _list, _db)
    # python test.py -db DB검색_2017_0703_17 -ds b5 -de b1100